import pandas as pd

"""
from openpyxl import load_workbook

# Carregar a planilha
workbook = load_workbook('planilha.xlsx')
worksheet = workbook.active  # Selecionando a primeira planilha

# Criar o arquivo TXT
with open('teste.txt', 'w') as f:
    for row in worksheet.iter_rows(values_only=True):
        line = '|'.join(str(cell) for cell in row)  # Unindo as células com tabulação
        f.write(line + '||\n')
"""

#Condicional para criar as planilhas
create = False

#Ler a planilha
df = pd.read_excel('../MovimentaçãoMercantis_ORIGINAL.xlsx')
df['CNPJ_ES'] = df['CNPJ_ES'].astype(str)
df['CPF_CNPJ'] = df['CPF_CNPJ'].apply(lambda x: str(x) if pd.notnull(x) else x)

# Método para aplicar função no DataFrame inteiro (apllymap)
#df = df.applymap(lambda x: str(x) if pd.notnull(x) else x)


desconto_devolucao = df[df['DESC_MERC_DEV'].apply(lambda x: float(x) != 0.0)] #Filtrando linhas
df = df[df['DESC_MERC_DEV'].apply(lambda x: float(x) == 0.0)] #Excluindo linhas filtradas
if (create):
    desconto_devolucao.to_excel('LINHAS_COM_DESC.POR_DEVOLUÇÃO_08.xlsx', index=False) #Criando tabela com as linhas filtradas

#DS_CC ou DS_BANCO
excluido_devolucao = df[df['DS_CC'].apply(lambda x: 'DEVOLUÇ' in x)] #Filtrando linhas
df = df[df['DS_CC'].apply(lambda x: not('DEVOLUÇ' in x))] #Excluindo linhas filtradas
if (create):
    excluido_devolucao.to_excel('LANCTO.EXCLUIDOS_EM_DEVOLUÇÃO_08.xlsx', index=False) #Criando tabela com as linhas filtradas

df_pagto = df[df['SAIDA'].apply(lambda x: x != 0 )] # Separando a tabela de PAGTO e RECEBTO
df_recebto = df[df['ENTRADA'].apply(lambda x: x != 0 )] # Separando a tabela de PAGTO e RECEBTO


# Removendo todos os espaços em branco de uma coluna
#df['Date'] = df['Date'].str.replace('\s+', '|', regex=True)

#Quebrando a coluna 'Date' onde tiver '|' e pegando apenas o último valor, e criando a coluna 'ano' para armazenar este registro
#df['ano'] = df['Date'].str.split('|').apply(lambda x: x[-1])

#Filtrando LINHAS à partir da coluna 'ano'
#df = df[df['ano'].apply(lambda x: int(x)) > 2020]

#Filtrando COLUNAS
#df = df[['Date', 'Open', 'Close']]  # Selecionar apenas essas colunas

#Removendo colunas pelo nome
df.drop(columns=['VALOR_PRINCIPAL', 'VALOR_REC_PAGO'], inplace=True)

print(df_pagto[['DATMOV', 'SAIDA', 'ENTRADA', 'NRO_NFE', 'HISTORICO']])
print(df_recebto[['DATMOV', 'SAIDA', 'ENTRADA', 'NRO_NFE', 'HISTORICO']])

#df.to_excel('Movimentação_Mercantis_08.xlsx', index=False)

"""
print(df.columns)
print("-----------------")
#Removendo colunas dinamicamente pelo nome
for column in df.columns:
    print(column)
    if column == 'VALOR_PRINCIPAL' or column == 'VALOR_REC_PAGO':
        df.drop(columns=[column], inplace=True)

print("-----------------")
print(df.columns)
"""

"""
# Salvando como CSV com personalizações
df.to_csv('teste.txt',
          sep='|',  # Delimitador: ponto e vírgula
          decimal=',',  # Separador decimal para números
          encoding='utf-8',  # Codificação
          index=False,  # Não incluir índice
          header=False)  # Incluir cabeçalho
"""